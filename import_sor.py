"""
ESPN SOR → Bracket Importer
============================
Updates SOS tab columns D-I from your ESPN BPI download.
Fixes the Excel date-conversion bug on W-L and QUAL WINS.

ONLY needs: openpyxl
  py -m pip install openpyxl

Usage:
  cd C:\\Users\\13032\\Downloads\\NCAAB
  python import_sor.py
"""

import datetime, sys, os

try:
    import openpyxl
except ImportError:
    print("Need openpyxl. Try:")
    print("  py -m pip install openpyxl")
    print("  python -m pip install openpyxl")
    sys.exit(1)

SOR_FILE = "ESPN_SOR.xlsx"
BRACKET_FILE = "ncaab_dynamic_bracket.xlsm"


def fix_date(val):
    """Excel converts '12-2' into Dec 2. Reverse it back to '12-2'."""
    if isinstance(val, datetime.datetime):
        return f"{val.month}-{val.day}", True
    if val is None or str(val).strip() == "":
        return "0-0", False
    return str(val).strip(), False


def split_wl(s):
    """Split '27-2' into (27, 2). First=wins, second=losses."""
    try:
        a, b = s.split("-")
        return int(a.strip()), int(b.strip())
    except:
        return 0, 0


def main():
    sor_file = SOR_FILE
    bracket_file = BRACKET_FILE

    # Parse command line args
    args = sys.argv[1:]
    i = 0
    while i < len(args):
        if args[i] == "--sor" and i + 1 < len(args):
            sor_file = args[i + 1]; i += 2
        elif args[i] == "--bracket" and i + 1 < len(args):
            bracket_file = args[i + 1]; i += 2
        else:
            i += 1

    if not os.path.exists(sor_file):
        print(f"ERROR: {sor_file} not found in {os.getcwd()}")
        sys.exit(1)
    if not os.path.exists(bracket_file):
        print(f"ERROR: {bracket_file} not found in {os.getcwd()}")
        sys.exit(1)

    print(f"ESPN file:   {sor_file}")
    print(f"Bracket:     {bracket_file}")
    print()

    # Read ESPN SOR
    espn = openpyxl.load_workbook(sor_file, data_only=True).active

    # ESPN columns:
    #   A(1)=Team  B(2)=CONF  C(3)=W-L  D(4)=SOR RK  E(5)=SOR SEED
    #   F(6)=SOR S-CURVE  G(7)=QUAL WINS  H(8)=SOS RK  I(9)=NC SOS
    print(f"ESPN headers: {[espn.cell(1,c).value for c in range(1,10)]}")

    # Load bracket (keep_vba=True for .xlsm)
    is_xlsm = bracket_file.lower().endswith('.xlsm')
    wb = openpyxl.load_workbook(bracket_file, keep_vba=is_xlsm)
    sos = wb["SOS"]

    # SOS tab columns:
    #   A(1)=Team  B(2)=CONF  C(3)=SOR RK(formula)  D(4)=SOS RK  E(5)=NC SOS
    #   F(6)=W     G(7)=L     H(8)=QW               I(9)=QL      J(10)=SOR Score(formula)
    print(f"SOS headers:  {[sos.cell(1,c).value for c in range(1,11)]}")
    print()

    # Build team lookup
    lookup = {}
    for r in range(2, sos.max_row + 1):
        t = sos.cell(r, 1).value
        if t:
            lookup[str(t).strip()] = r

    fixed_wl = 0
    fixed_qw = 0
    updated = 0
    skipped = 0

    for r in range(2, espn.max_row + 1):
        team = espn.cell(r, 1).value
        if not team:
            continue
        team = str(team).strip()

        if team not in lookup:
            skipped += 1
            continue

        row = lookup[team]
        conf = str(espn.cell(r, 2).value or "")

        # Read and fix W-L (ESPN col C)
        wl_str, was_date = fix_date(espn.cell(r, 3).value)
        if was_date:
            fixed_wl += 1
        w, l = split_wl(wl_str)

        # Read and fix QUAL WINS (ESPN col G)
        qw_str, was_date = fix_date(espn.cell(r, 7).value)
        if was_date:
            fixed_qw += 1
        qw, ql = split_wl(qw_str)

        # SOS RK (ESPN col H) and NC SOS (ESPN col I)
        sos_rk = espn.cell(r, 8).value
        nc_sos = espn.cell(r, 9).value

        # Write to SOS tab (only data columns, not formula columns C and J)
        sos.cell(row, 2).value = conf                              # B = CONF
        sos.cell(row, 4).value = int(sos_rk) if sos_rk else 200   # D = SOS RK
        sos.cell(row, 5).value = int(nc_sos) if nc_sos else 200   # E = NC SOS
        sos.cell(row, 6).value = w                                  # F = WINS
        sos.cell(row, 7).value = l                                  # G = LOSSES
        sos.cell(row, 8).value = qw                                 # H = QUAL WINS
        sos.cell(row, 9).value = ql                                 # I = QUAL LOSSES
        updated += 1

    wb.save(bracket_file)
    print(f"Updated {updated} teams")
    print(f"Skipped {skipped} teams (not in SOS tab)")
    print(f"W-L dates fixed:       {fixed_wl}")
    print(f"QUAL WINS dates fixed: {fixed_qw}")
    print()

    # Verify a few teams
    wb2 = openpyxl.load_workbook(bracket_file, data_only=True)
    ws2 = wb2["SOS"]
    print("Verification (first 5 teams):")
    for r in range(2, 7):
        t = ws2.cell(r, 1).value
        print(f"  {t}: W={ws2.cell(r,6).value} L={ws2.cell(r,7).value} "
              f"QW={ws2.cell(r,8).value} QL={ws2.cell(r,9).value} "
              f"SOS_RK={ws2.cell(r,4).value}")

    print(f"\nSaved to {bracket_file}")
    print("Open in Excel — SOR Score & SOR RK formulas will auto-recalculate.")


if __name__ == "__main__":
    main()
