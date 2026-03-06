"""
Build Dashboard — Reads bracket from Excel + EV bets from bot output.
Generates a single shareable dashboard.html with everything embedded.

Usage:
  python ncaab_bot.py          # generates output/latest.json
  python build_dashboard.py    # generates dashboard.html

Or just: python daily_push.py  (does both + pushes to GitHub)
"""
import json, sys, os, re
from pathlib import Path

def load_bracket(excel_path):
    """Read the Sim Engine tab and return bracket data as JSON-ready dicts."""
    try:
        import openpyxl
    except ImportError:
        print("Need openpyxl: py -m pip install openpyxl"); sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["⚙️ Sim Engine"]

    bracket = {}
    current_round = None

    # Row mapping from Sim Engine headers
    round_map = {
        "FIRST FOUR": "First Four",
        "ROUND OF 64": "Round of 64",
        "ROUND OF 32": "Round of 32",
        "SWEET 16": "Sweet 16",
        "ELITE 8": "Elite 8",
        "FINAL FOUR": "Final Four",
        "CHAMPIONSHIP": "Championship",
    }

    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if not val:
            continue
        val_str = str(val)

        # Detect section headers
        if "──" in val_str:
            for key, name in round_map.items():
                if key in val_str.upper():
                    current_round = name
                    bracket[current_round] = []
                    break
            continue

        # Skip non-game rows
        if not current_round:
            continue
        team_a = ws.cell(r, 3).value
        team_b = ws.cell(r, 5).value
        if not team_a or not team_b:
            continue

        score_a = ws.cell(r, 4).value
        score_b = ws.cell(r, 6).value
        spread = ws.cell(r, 7).value
        win_a = ws.cell(r, 8).value
        winner = ws.cell(r, 9).value

        if score_a is None or score_b is None:
            continue

        # Shorten team names for display
        def short(name):
            name = str(name)
            # Remove common suffixes
            for suf in [" Wildcats"," Blue Devils"," Wolverines"," Huskies"," Spartans",
                        " Gators"," Cougars"," Boilermakers"," Fighting Illini"," Crimson Tide",
                        " Tigers"," Cyclones"," Volunteers"," Golden Eagles"," Jayhawks",
                        " Red Storm"," Badgers"," Aggies"," Cardinals"," Ducks",
                        " Bulldogs"," Knights"," Commodores"," Flames"," Lumberjacks",
                        " RedHawks"," Midshipmen"," Bruins"," Buccaneers"," Warriors",
                        " Billikens"," Wolfpack"," Bulls"," Red Raiders"," Longhorns",
                        " Raiders"," Rainbow Warriors"," Bison"," Retrievers"," Broncos",
                        " Lobos"," Sharks"," Horned Frogs"," Hoosiers"," Hawkeyes",
                        " Razorbacks"," Cavaliers"," Panthers"," Buckeyes"," Cornhuskers",
                        " Trojans"," Governors"," Gaels"," Mustangs"," Vikings",
                        " Hurricanes"," Tar Heels"," Seahawks"," Phoenix"," Cougars",
                        " Bears"," Rams"," Miners"," Owls"," Pirates"]:
                if name.endswith(suf):
                    name = name[:-len(suf)]
                    break
            return name.strip()

        region = str(ws.cell(r, 2).value or "")

        game = {
            "r": region.split(" ")[0] if region else "",  # First word of region
            "a": short(team_a),
            "scA": round(float(score_a)),
            "b": short(team_b),
            "scB": round(float(score_b)),
            "sp": round(float(spread), 1),
            "w": round(float(win_a), 3),
            "win": short(winner) if winner else "",
        }

        # Add seed info for Round of 64 and First Four
        # Parse from Bracket Input tab if possible, otherwise skip
        bracket[current_round].append(game)

    wb.close()

    # Try to add seeds from Bracket Input
    try:
        wb2 = openpyxl.load_workbook(excel_path, data_only=True)
        ws2 = wb2["📋 Bracket Input"]
        seeds = {}  # team_short_name → seed
        for r in range(5, ws2.max_row + 1):
            seed = ws2.cell(r, 3).value
            team = ws2.cell(r, 4).value
            if seed and team and isinstance(seed, (int, float)):
                seeds[short(str(team))] = int(seed)
            seed2 = ws2.cell(r, 6).value
            team2 = ws2.cell(r, 7).value
            if seed2 and team2 and isinstance(seed2, (int, float)):
                s = str(team2)
                if "Winner" not in s:
                    seeds[short(s)] = int(seed2)
        wb2.close()

        # Apply seeds to First Four and Round of 64
        for rnd in ["First Four", "Round of 64"]:
            if rnd in bracket:
                for g in bracket[rnd]:
                    if g["a"] in seeds:
                        g["sA"] = seeds[g["a"]]
                    if g["b"] in seeds:
                        g["sB"] = seeds[g["b"]]
    except:
        pass  # Seeds are nice-to-have, not critical

    # Determine champion and final four
    champ = bracket.get("Championship", [{}])[0].get("win", "Duke")
    ff_teams = [g["win"] for g in bracket.get("Elite 8", [])]

    # Region champions
    region_champs = {}
    for g in bracket.get("Elite 8", []):
        region_champs[g["r"]] = g["win"]

    return bracket, champ, ff_teams, region_champs


def load_bot_data():
    for p in ["output/latest.json", "latest.json"]:
        if os.path.exists(p):
            with open(p) as f:
                return json.load(f)
    return None


def main():
    # Find Excel file
    excel = None
    for name in ["ncaab_dynamic_bracket.xlsm", "ncaab_dynamic_bracket.xlsx"]:
        if os.path.exists(name):
            excel = name
            break

    if not excel:
        print("ERROR: No bracket Excel file found")
        sys.exit(1)

    print(f"Reading bracket from {excel}...")
    bracket, champ, ff_teams, region_champs = load_bracket(excel)
    for rnd, games in bracket.items():
        print(f"  {rnd}: {len(games)} games")

    bot_data = load_bot_data()
    if bot_data:
        print(f"Bot data: {bot_data.get('date')} — {len(bot_data.get('ev_bets',[]))} EV bets")
    else:
        print("No bot data found (output/latest.json)")

    bracket_json = json.dumps(bracket, default=str)
    bot_json = json.dumps(bot_data, default=str) if bot_data else "null"
    champ_json = json.dumps(champ)
    ff_json = json.dumps(ff_teams)
    rc_json = json.dumps(region_champs)

    html = TEMPLATE.replace("__BRACKET__", bracket_json)
    html = html.replace("__BOTDATA__", bot_json)
    html = html.replace("__CHAMP__", champ_json)
    html = html.replace("__FF__", ff_json)
    html = html.replace("__RC__", rc_json)

    out = Path("dashboard.html")
    with open(out, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Dashboard: {out} ({out.stat().st_size:,} bytes)")


TEMPLATE = r'''<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>NCAAB Bracket + EV Dashboard</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@300;400;600;700;800&display=swap');
*{margin:0;padding:0;box-sizing:border-box}body{font-family:'JetBrains Mono',monospace;background:#0a0f1a;color:#e2e8f0;min-height:100vh}
.hd{background:linear-gradient(135deg,#0f172a,#1e1b4b,#0f172a);border-bottom:1px solid #1e293b;padding:16px 24px 0}
.tl{display:flex;align-items:baseline;gap:12px;flex-wrap:wrap;margin-bottom:4px}
.tl h1{font-size:22px;font-weight:800;letter-spacing:-1px}.tl .su{font-size:22px;font-weight:300;color:#818cf8}
.tl .ch{font-size:11px;color:#fbbf24;margin-left:auto;font-weight:700}
.mt{display:flex;gap:10px;font-size:10px;color:#64748b;margin-bottom:12px;flex-wrap:wrap}
.tabs{display:flex;gap:0;overflow-x:auto}
.tab{padding:10px 18px;font-size:12px;font-weight:400;color:#64748b;background:0;border:none;border-bottom:2px solid transparent;cursor:pointer;font-family:inherit;border-radius:6px 6px 0 0;white-space:nowrap}
.tab.on{font-weight:700;color:#f8fafc;background:#1e293b;border-bottom-color:#818cf8}
.ct{padding:20px 24px}.hid{display:none!important}
.pls{display:flex;gap:4px;margin-bottom:10px;flex-wrap:wrap}
.pl{padding:5px 12px;font-size:11px;color:#94a3b8;background:#1e293b;border:1px solid #334155;border-radius:20px;cursor:pointer;font-family:inherit}
.pl.on{font-weight:700;color:#f8fafc;background:#818cf8;border-color:#818cf8}
.rgs{display:flex;gap:4px;margin-bottom:14px}
.rb{padding:3px 10px;font-size:10px;background:0;border:1px solid #334155;border-radius:4px;cursor:pointer;font-family:inherit}
.gm{background:#111827;border:1px solid #1e293b;border-radius:6px;padding:8px 12px;font-size:12px;margin-bottom:4px;border-left:3px solid #475569}
.gm .rw{display:flex;justify-content:space-between;padding:1px 0}
.gm .w{color:#f8fafc;font-weight:700}.gm .l{color:#64748b}
.gm .sd{color:#475569;font-size:9px;margin-right:4px}.gm .sc{color:#94a3b8;min-width:30px;text-align:right;font-weight:600}
.gm .mr{display:flex;justify-content:space-between;font-size:9px;color:#475569;margin-top:3px;padding-top:3px;border-top:1px solid #1e293b}
.gm .up{color:#fbbf24;font-weight:700}
.gm.EAST{border-left-color:#3b82f6}.gm.SOUTH{border-left-color:#f97316}.gm.MIDWEST{border-left-color:#10b981}.gm.WEST{border-left-color:#a855f7}
.gg{display:grid;gap:6px}.g4{grid-template-columns:repeat(4,1fr)}.g2{grid-template-columns:1fr 1fr}.g1{grid-template-columns:1fr}
.pt{margin-top:16px;background:#111827;border:1px solid #1e293b;border-radius:8px;padding:16px}
.pt h3{font-size:12px;font-weight:700;color:#818cf8;margin-bottom:8px}
.pg{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;font-size:11px}
.pc{padding-top:6px}.pc .rn{font-weight:700;font-size:10px;text-transform:uppercase;letter-spacing:1px}
.pc .cn{color:#f8fafc;font-weight:800;font-size:14px;margin:2px 0}.pc .tr{color:#475569;font-size:9px}
.bt{background:#111827;border:1px solid #1e293b;border-radius:8px;padding:14px;margin-bottom:8px;display:grid;grid-template-columns:1fr auto 1fr;gap:16px;align-items:center}
.bt.hot{border-left:3px solid #16a34a}.bt.warm{border-left:3px solid #ca8a04}.bt.cool{border-left:3px solid #6b7280}
.bt .lb{font-size:10px;color:#64748b;text-transform:uppercase;letter-spacing:1px}
.bt .si{font-size:14px;font-weight:700;margin-top:2px}.bt .gn{font-size:11px;color:#94a3b8;margin-top:2px}
.bt .dt{font-size:10px;color:#475569;margin-top:2px}.bt .od{font-size:24px;font-weight:800;text-align:center}
.bt .li{font-size:10px;color:#94a3b8;text-align:center}
.eb{display:inline-block;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:700;color:#fff}
.eh{background:#16a34a}.ew{background:#ca8a04}.ec{background:#6b7280}
.pb{position:relative;height:8px;background:#1e293b;border-radius:4px;overflow:hidden;margin-top:4px}
.pi{position:absolute;left:0;top:0;height:100%;background:#475569;border-radius:4px}
.pm{position:absolute;left:0;top:0;height:100%;background:#4ade80;border-radius:4px;opacity:.7}
table{width:100%;border-collapse:collapse;font-size:12px}
th{padding:8px 6px;color:#64748b;font-weight:600;font-size:10px;cursor:pointer;user-select:none;border-bottom:2px solid #1e293b}
th.sr{color:#818cf8}td{padding:7px 6px;border-bottom:1px solid #1e293b}tr:nth-child(even){background:#0d1321}
.pn{background:#111827;border:1px solid #1e293b;border-radius:8px;padding:20px;margin-bottom:16px}
.pn h3{font-size:14px;font-weight:700;color:#818cf8;margin-bottom:12px}
.tc{display:grid;grid-template-columns:1fr 1fr;gap:16px}
.sb{text-align:center;padding:12px;background:#0a0f1a;border-radius:6px}
.sb .n{font-size:28px;font-weight:800}.sb .la{font-size:11px;color:#64748b;margin-top:4px}.sb .sm{font-size:10px;color:#475569}
.ff{background:linear-gradient(135deg,#1e1b4b,#172554);border:1px solid #3730a3;border-radius:8px;padding:16px;text-align:center}
.ff .fl{font-size:10px;color:#818cf8;text-transform:uppercase;letter-spacing:2px;margin-bottom:12px}
.ff .fc{margin-top:12px;font-size:18px;font-weight:800;color:#fbbf24}
@media(max-width:768px){.g4{grid-template-columns:1fr 1fr}.pg{grid-template-columns:1fr 1fr}.tc{grid-template-columns:1fr}.bt{grid-template-columns:1fr;gap:8px}}
</style></head><body>
<div class="hd">
<div class="tl"><h1>NCAAB</h1><span class="su">BRACKET + EV</span><span class="ch" id="chLbl"></span></div>
<div class="mt"><span style="color:#4ade80">74.98% SU</span><span>•</span><span style="color:#fbbf24">56.9% ATS</span><span>•</span><span id="ffLbl"></span><span>•</span><span id="ds"></span></div>
<div class="tabs"><button class="tab on" onclick="st('bracket',this)">Bracket</button><button class="tab" onclick="st('ev',this)">EV Bets</button><button class="tab" onclick="st('rankings',this)">Rankings</button><button class="tab" onclick="st('model',this)">Model</button></div>
</div>
<div class="ct"><div id="t-bracket"></div><div id="t-ev" class="hid"></div><div id="t-rankings" class="hid"></div><div id="t-model" class="hid"></div></div>
<script>
const B=__BRACKET__;
const bd=__BOTDATA__;
const CHAMP=__CHAMP__;
const FF=__FF__;
const RCHAMPS=__RC__;
const RC={EAST:"#3b82f6",SOUTH:"#f97316",MIDWEST:"#10b981",WEST:"#a855f7"};
const RN=Object.keys(B);

document.getElementById("chLbl").textContent="🏆 PROJECTED: "+CHAMP;
document.getElementById("ffLbl").textContent="F4: "+FF.join(" | ");
if(bd){document.getElementById("ds").innerHTML='<span style="color:#4ade80">● '+bd.date+' — '+(bd.ev_bets||[]).length+' EV bets</span>'}
else{document.getElementById("ds").innerHTML='<span style="color:#fbbf24">○ No bot data</span>'}

let cr=RN.length>1?RN[1]:RN[0],cg="ALL",sc="rk",sd=1;
function st(id,el){["bracket","ev","rankings","model"].forEach(t=>document.getElementById("t-"+t).classList.add("hid"));document.getElementById("t-"+id).classList.remove("hid");document.querySelectorAll(".tab").forEach(t=>t.classList.remove("on"));el.classList.add("on")}

function gc(g){
  const u=g.w<.5;
  const wA=g.win===g.a, wB=g.win===g.b;
  return '<div class="gm '+g.r+'">'+
    '<div class="rw"><span class="'+(wA?'w':'l')+'">'+(g.sA!=null?'<span class="sd">'+g.sA+'</span>':'')+g.a+'</span><span class="sc">'+g.scA+'</span></div>'+
    '<div class="rw"><span class="'+(wB?'w':'l')+'">'+(g.sB!=null?'<span class="sd">'+g.sB+'</span>':'')+g.b+'</span><span class="sc">'+g.scB+'</span></div>'+
    '<div class="mr"><span>'+(g.sp>0?'+':'')+g.sp.toFixed(1)+'</span><span>'+(g.w*100).toFixed(0)+'–'+((1-g.w)*100).toFixed(0)+'%</span>'+(u?'<span class="up">UPSET</span>':'')+'</div></div>';
}

function rb(){
  let gm=B[cr]||[];
  if(cg!=="ALL")gm=gm.filter(g=>g.r===cg||g.r.includes(cg.substring(0,2)));
  const p=RN.map(r=>'<button class="pl '+(r===cr?'on':'')+'" onclick="cr=\''+r+'\';rb()">'+r+'</button>').join('');
  const rg=["ALL","EAST","SOUTH","MIDWEST","WEST"].map(r=>'<button class="rb '+(r===cg?'on':'')+'" style="color:'+(r===cg?'#f8fafc':(RC[r]||'#94a3b8'))+';'+(r===cg?'background:'+(RC[r]||'#475569')+';':'')+'border-color:'+(RC[r]||'#334155')+'" onclick="cg=\''+r+'\';rb()">'+r+'</button>').join('');
  let body='';
  if(cr==="Final Four"||cr==="Championship"){
    body='<div class="ff"><div class="fl">'+(cr==="Championship"?"NATIONAL CHAMPIONSHIP":"FINAL FOUR")+'</div>'+gm.map(gc).join('')+(cr==="Championship"?'<div class="fc">🏆 '+CHAMP+' — Projected Champions</div>':'')+'</div>';
  }else{
    const cl=gm.length>8?'g4':gm.length>4?'g2':'g1';
    body='<div class="gg '+cl+'">'+gm.map(gc).join('')+'</div>';
  }
  // Region champs footer
  const rcHTML=Object.entries(RCHAMPS).map(([r,t])=>'<div class="pc" style="border-top:2px solid '+(RC[r]||'#475569')+'"><div class="rn" style="color:'+(RC[r]||'#94a3b8')+'">'+r+'</div><div class="cn">'+t+'</div></div>').join('');
  document.getElementById("t-bracket").innerHTML='<div class="pls">'+p+'</div><div class="rgs">'+rg+'</div>'+body+'<div class="pt"><h3>Region Champions</h3><div class="pg">'+rcHTML+'</div></div>';
}

function rev(){
  const el=document.getElementById("t-ev");
  if(!bd||!bd.ev_bets||!bd.ev_bets.length){el.innerHTML='<p style="color:#94a3b8">No EV bets. Run: <code style="color:#4ade80">python ncaab_bot.py</code> then <code style="color:#4ade80">python build_dashboard.py</code></p>';return}
  el.innerHTML='<div style="font-size:12px;color:#94a3b8;margin-bottom:14px">'+bd.ev_bets.length+' +EV bets — '+bd.date+'</div>'+
    bd.ev_bets.slice(0,30).map(function(b){
      var e=b.ev_pct>=6?'hot':b.ev_pct>=4?'warm':'cool';
      var o=b.odds>0?'+'+b.odds:''+b.odds;
      var ed=((b.model_prob-b.implied)*100).toFixed(1);
      return '<div class="bt '+e+'"><div><div class="lb">'+b.type+' • '+b.book+'</div><div class="si">'+b.side+'</div><div class="gn">'+b.game+'</div><div class="dt">Pred: '+b.pred+' • Spread: '+(b.model_spread>0?'+':'')+b.model_spread+'</div></div><div style="text-align:center"><div class="od">'+o+'</div>'+(b.line!=null?'<div class="li">Line: '+(b.line>0?'+':'')+b.line+'</div>':'')+'<span class="eb e'+e[0]+'">+'+b.ev_pct+'% EV</span></div><div><div style="display:flex;justify-content:space-between;font-size:10px;color:#64748b;margin-bottom:2px"><span>Probability</span><span style="color:#4ade80">Edge: +'+ed+'%</span></div><div class="pb"><div class="pi" style="width:'+b.implied*100+'%"></div><div class="pm" style="width:'+b.model_prob*100+'%"></div></div><div style="display:flex;justify-content:space-between;font-size:9px;color:#475569;margin-top:1px"><span>Model: '+(b.model_prob*100).toFixed(0)+'%</span><span>Market: '+(b.implied*100).toFixed(0)+'%</span></div></div></div>';
    }).join('');
}

function rrk(){
  var T=bd&&bd.top_25?bd.top_25.map(function(t,i){return{rk:t.rank||i+1,nm:t.name,cf:t.conf||'',rc:t.record,sr:t.sor,ss:t.sos||0,po:t.ppp_off,pd:t.ppp_def,q:t.qual}}):[{rk:1,nm:"No data",cf:"",rc:"",sr:0,ss:0,po:0,pd:0,q:""}];
  var s=T.slice().sort(function(a,b){var av=a[sc],bv=b[sc];return typeof av==='number'?(av-bv)*sd:String(av||'').localeCompare(String(bv||''))*sd});
  var cols=[{k:'rk',l:'#'},{k:'nm',l:'TEAM'},{k:'cf',l:'CONF'},{k:'rc',l:'REC'},{k:'sr',l:'SOR'},{k:'ss',l:'SOS'},{k:'po',l:'PPP↑'},{k:'pd',l:'PPP↓'},{k:'q',l:'QUAL'}];
  document.getElementById("t-rankings").innerHTML='<div style="overflow-x:auto"><table><thead><tr>'+cols.map(function(c){return'<th class="'+(sc===c.k?'sr':'')+'" style="text-align:'+(c.k==='nm'?'left':'center')+'" onclick="sc=\''+c.k+'\';sd=sc===\''+c.k+'\'?-sd:1;rrk()">'+c.l+(sc===c.k?(sd===1?' ▲':' ▼'):'')+'</th>'}).join('')+'</tr></thead><tbody>'+s.map(function(t){return'<tr><td style="text-align:center;font-weight:800;color:'+(t.rk<=4?'#818cf8':t.rk<=10?'#f8fafc':'#94a3b8')+'">'+t.rk+'</td><td style="font-weight:600">'+t.nm+'</td><td style="text-align:center;color:#64748b;font-size:10px">'+t.cf+'</td><td style="text-align:center">'+t.rc+'</td><td style="text-align:center;font-weight:700;color:'+(t.sr>=.7?'#4ade80':t.sr>=.6?'#fbbf24':'#94a3b8')+'">'+t.sr.toFixed(3)+'</td><td style="text-align:center">'+t.ss+'</td><td style="text-align:center;color:'+(t.po>=1.2?'#4ade80':'#e2e8f0')+'">'+t.po.toFixed(3)+'</td><td style="text-align:center;color:'+(t.pd<=.97?'#4ade80':t.pd>=1.05?'#f87171':'#e2e8f0')+'">'+t.pd.toFixed(3)+'</td><td style="text-align:center">'+t.q+'</td></tr>'}).join('')+'</tbody></table></div>';
}

document.getElementById("t-model").innerHTML='<div class="tc"><div class="pn"><h3>Model Architecture</h3><div style="font-size:12px;line-height:1.8;color:#94a3b8"><div><span style="color:#4ade80">→</span> Blended PPP = Off_w × PPP_off(adj) + Def_w × PPP_def(adj)</div><div><span style="color:#4ade80">→</span> SOS Adj = PPP × ((366-SOS_RK)/365)^α</div><div><span style="color:#4ade80">→</span> Pace = 0.65×min(A,B) + 0.35×max(A,B)</div><div><span style="color:#4ade80">→</span> Score = PPP×Pace + HCA + SOR_adj + Foul</div><div><span style="color:#4ade80">→</span> Win% = 1/(1+e^(-spread/18))</div></div></div><div class="pn"><h3>Backtest</h3><div class="tc"><div class="sb"><div class="n" style="color:#4ade80">74.98%</div><div class="la">Straight Up</div><div class="sm">3,290 games</div></div><div class="sb"><div class="n" style="color:#fbbf24">56.94%</div><div class="la">ATS (5pt edge)</div><div class="sm">353 bets</div></div></div><div style="margin-top:12px;font-size:12px;color:#94a3b8"><b style="color:#f8fafc">Params:</b> Off=0.44 Def=0.43 SOR=0.03 HCA=1 Scale=18</div></div></div>';
rb();rev();rrk();
</script></body></html>'''


if __name__ == "__main__":
    main()
